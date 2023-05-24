import openpyxl

def get_unique_funds(sheet, row):
    unique_funds = {}

    for col in range(6, sheet.max_column + 1):
        cell_value = sheet.cell(row=row, column=col).value
        if cell_value is not None:
            if cell_value not in unique_funds:
                unique_funds[cell_value] = []
            unique_funds[cell_value].append(sheet.title)
            print(f"Extracted fund: {cell_value} from sheet: {sheet.title}")

    return unique_funds

def write_unique_funds_to_excel(worksheet, row, unique_funds, header):
    worksheet.cell(row=row, column=1, value=header)

    for index, (fund, sheets) in enumerate(unique_funds.items(), start=row):
        worksheet.cell(row=index, column=2, value=fund)
        worksheet.cell(row=index, column=3, value=', '.join(sheets))


years = ["2021matticopy", "2020", "2019", "2018", "2017", "2016", "2015", "2014", "2013", "2012", "2011", "2010", "2009", "2008", "2007", "2006", "2005", "2004", "2003", "2002", "2001", "2000", "1999", "1998", "1997"]


unique_funds = {}

for year in years:
    file_path = f'/users/matthias/Documents/GitHub/engx-project-group20/skjölin frá Birgi/Files/Arsreikningar-lifeyrissjoda_{year}.xlsx'
    workbook = openpyxl.load_workbook(file_path)

    if year != "2021matticopy":
        if int(year) <= 2010:
            sheets = [workbook["funddata sam"], workbook["funddata ser"]]
        else:
            sheets = [workbook["funddata sam"], workbook["funddata ser"], workbook["funddata hluti 3"]]
    else:
        sheets = [workbook["funddata sam"], workbook["funddata ser"], workbook["funddata hluti 3"]]

    for sheet in sheets:
        unique_funds.update(get_unique_funds(sheet, row=1))  # Change row number to 1

unique_funds = dict(sorted(unique_funds.items()))

# Create a new Excel workbook and sheet for output
output_workbook = openpyxl.Workbook()
output_sheet = output_workbook.active
output_sheet.title = "Unique Funds"

write_unique_funds_to_excel(output_sheet, row=1, unique_funds=unique_funds, header="Funds")

# Save the output workbook
output_workbook.save("unique_fundsV2.xlsx")