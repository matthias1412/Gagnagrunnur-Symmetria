import openpyxl

def get_unique_subfunds(sheet, row):
    unique_subfunds = {}

    for col in range(6, sheet.max_column + 0):
        cell_value = sheet.cell(row=row, column=col).value
        if cell_value is not None:
            if cell_value not in unique_subfunds:
                unique_subfunds[cell_value] = []
            unique_subfunds[cell_value].append(sheet.title)
            print(f"Extracted subfund: {cell_value} from sheet: {sheet.title}")


    return unique_subfunds

def write_unique_subfunds_to_excel(worksheet, row, unique_subfunds, header):
    worksheet.cell(row=row, column=1, value=header)

    for index, (subfund, sheets) in enumerate(unique_subfunds.items(), start=row):
        worksheet.cell(row=index, column=2, value=subfund)
        worksheet.cell(row=index, column=3, value=', '.join(sheets))

years = ["2021matticopy", "2020", "2019", "2018", "2017", "2016", "2015", "2014", "2013", "2012", "2011", "2010", "2009", "2008", "2007", "2006", "2005", "2004", "2003", "2002", "2001", "2000", "1999", "1998", "1997"]


unique_subfunds = {}

for year in years:
    file_path = f'/users/matthias/Documents/GitHub/engx-project-group20/skjölin frá Birgi/Files/Arsreikningar-lifeyrissjoda_{year}.xlsx'
    workbook = openpyxl.load_workbook(file_path)

    if year != "2021matticopy":
        if int(year) <= 2010:
            sheets = [workbook["funddata sam"], workbook["funddata ser"]]
        else:
            sheets = [workbook["funddata sam"], workbook["funddata ser"], workbook["funddata hluti 3"]]
    else:
        sheets = [workbook["funddata sam"], workbook["funddata ser"], workbook["funddata hluti 3"]]

    for sheet in sheets:
        unique_subfunds.update(get_unique_subfunds(sheet, row=2))

unique_subfunds = dict(sorted(unique_subfunds.items()))

# Create a new Excel workbook and sheet for output
output_workbook = openpyxl.Workbook(write_only=True)
output_sheet = output_workbook.create_sheet("Unique Subfunds")



write_unique_subfunds_to_excel(output_sheet, row=1, unique_subfunds=unique_subfunds, header="Subfunds")

# Save the output workbook
with open("unique_fundsV2.xlsx", "wb") as output_file:
    output_workbook.save(output_file)

