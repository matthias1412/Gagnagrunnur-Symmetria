import openpyxl

def get_unique_values(sheet, column, year):
    unique_values = {}

    for row in range(4, sheet.max_row + 1):
        cell_value = sheet.cell(row=row, column=column).value
        if cell_value is not None:
            if cell_value not in unique_values:
                unique_values[cell_value] = []
            unique_values[cell_value].append((year, sheet.title))

    return unique_values

def write_unique_values_to_excel(worksheet, row, unique_values, header):
    worksheet.cell(row=row, column=1, value=header)

    for index, (value, locations) in enumerate(unique_values.items(), start=row):
        worksheet.cell(row=index, column=2, value=value)
        worksheet.cell(row=index, column=3, value=', '.join(f'{year}-{sheet}' for year, sheet in locations))

years = ["2021matticopy", "2020", "2019", "2018", "2017", "2016", "2015", "2014", "2013", "2012", "2011","2010","2009","2008","2007","2006","2005","2004","2003","2002","2001","2000","1999","1998","1997"]

# Initialize dictionaries for unique attribute names
unique_attribute1_names = {}
unique_attribute2_names = {}
unique_attribute3_names = {}
unique_attribute4_names = {}

for year in years:
    
    # Load the Excel workbook and sheets
    file_path = f'/users/matthias/Documents/GitHub/engx-project-group20/skjölin frá Birgi/Files/Arsreikningar-lifeyrissjoda_{year}.xlsx'
    workbook = openpyxl.load_workbook(file_path)

    if year != "2021matticopy":
        if int(year) <= 2010:
            sheet1 = workbook["funddata sam"]
            sheet2 = workbook["funddata ser"]
            sheets = [sheet1, sheet2]

        else:
            sheet1 = workbook["funddata sam"]
            sheet2 = workbook["funddata ser"]
            sheet3 = workbook["funddata hluti 3"]
            sheets = [sheet1, sheet2, sheet3]
    else: 
            sheet1 = workbook["funddata sam"]
            sheet2 = workbook["funddata ser"]
            sheet3 = workbook["funddata hluti 3"]
            sheets = [sheet1, sheet2, sheet3]


    # Extract unique attribute names for each attribute column
    for sheet in sheets:
        unique_attribute1_names.update(get_unique_values(sheet, column=1, year=year))
        unique_attribute2_names.update(get_unique_values(sheet, column=2, year=year))
        unique_attribute3_names.update(get_unique_values(sheet, column=3, year=year))
        unique_attribute4_names.update(get_unique_values(sheet, column=4, year=year))

# Sort the dictionaries by key
unique_attribute1_names = dict(sorted(unique_attribute1_names.items()))
unique_attribute2_names = dict(sorted(unique_attribute2_names.items()))
unique_attribute3_names = dict(sorted(unique_attribute3_names.items()))
unique_attribute4_names = dict(sorted(unique_attribute4_names.items()))

# Create a new Excel workbook and sheet for output
output_workbook = openpyxl.Workbook()
output_sheet = output_workbook.active
output_sheet.title = "Unique Attribute Values"

# Write unique attribute names to the output sheet
write_unique_values_to_excel(output_sheet, row=1, unique_values=unique_attribute1_names, header="Attribute 1")
write_unique_values_to_excel(output_sheet, row=len(unique_attribute1_names) + 2, unique_values=unique_attribute2_names, header="Attribute 2")
write_unique_values_to_excel(output_sheet, row=len(unique_attribute1_names) + len(unique_attribute2_names) + 3, unique_values=unique_attribute3_names, header="Attribute 3")
write_unique_values_to_excel(output_sheet, row=len(unique_attribute1_names) + len(unique_attribute2_names) + len(unique_attribute3_names) + 4, unique_values=unique_attribute4_names, header="Attribute 4")

# Save the output workbook
output_workbook.save("unique_attribute_valuesV2.xlsx")

